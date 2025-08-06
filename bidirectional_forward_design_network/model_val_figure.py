import os
import sys

import openpyxl
import torch
import torch.nn as nn
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_absolute_error
from torch.utils.data import DataLoader

import auxiliary_tools.parseUnit as parseUtils
from forward_design_network import ForwardNetMlp as forward_design_net
from auxiliary_tools.dataSet_forward import DataSet

args = parseUtils.MyParse(debug=False).args

# 加载数据集
dataset_tar = torch.load(r"D:\DL\Valley_Inverse_Design\tools\dataset_tar\unloaded_dataset_20018_forward.pth")
dataset_train = dataset_tar['dataset_train']
dataset_val = dataset_tar['dataset_val']
train_loader = DataLoader(dataset=dataset_train, batch_size=16, shuffle=True)
val_loader = DataLoader(dataset=dataset_val, batch_size=64, shuffle=True)

# 加载模型
model = forward_design_net().to(args.device)
forward_tar = r"../data_space/checkpoints/bidirectional-forward-design-network-checkpoint.pth.tar"

if os.path.exists(forward_tar):
    print("load forward checkpoint...")
    checkpoint = torch.load(forward_tar)
    model.load_state_dict(checkpoint['state_dict'])
    print(checkpoint['epoch'])
else:
    print("no found")


# 评估模型函数
def evaluate_model(loader, model, device):
    model.eval()
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for structure_nor, valley_pos, _, _ in loader:
            structure = structure_nor.to(device)
            valley_pos = valley_pos.to(device)

            # 正向传播
            predict_valley = model(structure)

            # 保存预测值和标签
            all_preds.append(predict_valley.cpu().numpy())
            all_labels.append(valley_pos.cpu().numpy())

    # 合并所有预测和标签
    preds = np.concatenate(all_preds, axis=0)
    labels = np.concatenate(all_labels, axis=0)

    return preds, labels


# 从训练集和验证集中获取预测值和真实值
train_preds, train_labels = evaluate_model(train_loader, model, args.device)
val_preds, val_labels = evaluate_model(val_loader, model, args.device)

# 设置随机种子并抽取1000个样本
np.random.seed(190730)
train_indices = np.random.choice(len(train_preds), 1000, replace=False)
val_indices = np.random.choice(len(val_preds), 1700, replace=False)

# 获取抽样后的数据
train_preds_sample = train_preds[train_indices]
train_labels_sample = train_labels[train_indices]
val_preds_sample = val_preds[val_indices]
val_labels_sample = val_labels[val_indices]

# 绘制散点图
plt.figure(figsize=(10, 6))

# # 创建一个新的工作簿
# wb = openpyxl.Workbook()
# # 选择活动的工作表
# ws = wb.active
# # 在第一列第一格写入字符串a
# ws.cell(row=1, column=1, value="pred")
# ws.cell(row=1, column=2, value="label")
# val_preds_sample_list = val_preds_sample.tolist()
# val_labels_sample_list = val_labels_sample.tolist()
# # 从第二行开始写入列表b中的数据
# for idx, value in enumerate(val_preds_sample, start=2):  # 从第二行开始
#     ws.cell(row=idx, column=1, value=float(value))
# for idx, value in enumerate(val_labels_sample, start=2):  # 从第二行开始
#     ws.cell(row=idx, column=2, value=float(value))
# # 指定保存的路径和文件名
# filename = f"RFNN_pred_vs_label"
# file_path = f"E:/桌面/{filename}.xlsx"
# wb.save(file_path)  # 保存Excel文件
# # 关闭工作簿
# wb.close()

# 验证集散点图 - 实心蓝色正方形
plt.xlim(0, 1)
plt.ylim(0, 1)
plt.scatter(val_preds_sample, val_labels_sample, color='blue', marker='s', label='Validation Set', alpha=0.4, s=50)
# 训练集散点图 - 红色圆点
# plt.scatter(train_preds_sample, train_labels_sample, color='red', marker='o', label='Training Set', alpha=0.4, s=50)



# 图例和标签
plt.xlabel('Predicted Value')
plt.ylabel('True Value')
plt.legend()
plt.title('Predicted vs True Values for Training and Validation Sets')
plt.show()
