import numpy as np
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer
from matplotlib import font_manager

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, \
    NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

import sys
from structure_design_form import Ui_Form

import os
import math
import torch
import numpy as np
from scipy.ndimage import gaussian_filter1d
import openpyxl

import auxiliary_tools.parseUnit as parseUtils
from bidirectional_forward_design_network.forward_design_network import ForwardNetMlp as forward_design_net
from spectrum_unloaded_prediction_network.spectrum_prediction_network import ForwardNetMlp as forward2_design_net
from bidirectional_inverse_design_network.inverse_design_network import InverseNet as inverse_design_net

import matplotlib.pyplot as plt

# 设置全局字体为 Arial
plt.rcParams['font.family'] = 'Arial'

args = parseUtils.MyParse(debug=False).args
args.device = "cpu"
i_model = inverse_design_net().to(args.device)
f_model = forward_design_net().to(args.device)
f2_model = forward2_design_net().to(args.device)

forward_tar = r"../data_space/checkpoints/bidirectional-forward-design-network-checkpoint.pth.tar"
if os.path.exists(forward_tar):
    print("load forward checkpoint...")
    f_checkpoint = torch.load(forward_tar)
    f_model.load_state_dict(f_checkpoint['state_dict'])
    print(f_checkpoint['epoch'])
else:
    print("no found")
forward2_tar = r"../data_space/checkpoints/spectrum-unloaded-prediction-network-checkpoint.pth.tar"
if os.path.exists(forward2_tar):
    print("load forward checkpoint...")
    f2_checkpoint = torch.load(forward2_tar)
    f2_model.load_state_dict(f2_checkpoint['state_dict'])
    print(f2_checkpoint['epoch'])
else:
    print("no found")

inverse_tar = r"../data_space/checkpoints/bidirectional-inverse-design-network-checkpoint.pth.tar"
if os.path.exists(inverse_tar):
    print("load inverse checkpoint_inverse.pth...")
    checkpoint = torch.load(inverse_tar)
    i_model.load_state_dict(checkpoint['state_dict'])
    print(checkpoint['epoch'])
else:
    print("no found")


def data_trans_inverse(data):
    data = data.view(1, -1)
    scale = args.structure_range
    for i in range(0, data.shape[1]):
        data[0][i] = data[0][i] * (scale[i][1] - scale[i][0]) + scale[i][0]
    return data


def data_trans(data):
    scale = args.structure_range
    for i in range(0, data.__len__()):
        data[i] = (data[i] - scale[i][0]) / (scale[i][1] - scale[i][0])
    return data


global point, ims
global button_press_flag, nearly_point_index, update_nearlyPoint_flag
global true_spectrum, predict_spectrum, predict_structure
button_press_flag = 0
nearly_point_index = 0
update_nearlyPoint_flag = 0
global point  # 可移动的点，根据这些点的位置插值
point = [1.0, 0.5]
res = []

class Mytest(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(Mytest, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Structure_design")
        self.ClearButton.clicked.connect(self.clear_predict)
        self.GetButton.clicked.connect(self.get_predict)
        self.init()

    def init(self, toolbarVisible=False, showHint=False):
        # excel相关
        self.counter = 0
        self.spectrum_predict = []
        self.structure_predict = []
        # 创建一个新的工作簿
        self.wb = openpyxl.Workbook()

        self.lineEdit_d1.setText('16')
        self.lineEdit_d2.setText('6')
        self.lineEdit_gx.setText('6')
        self.lineEdit_gy.setText('PV')

        self.static_canvas = FigureCanvas(Figure())
        self.naviBar = NavigationToolbar(self.static_canvas, self)
        actList = self.naviBar.actions()
        count = len(actList)
        self.__lastActtionHint = actList[count - 1]
        self.__showHint = showHint  # 是否显示坐标提示
        self.__lastActtionHint.setVisible(self.__showHint)
        self.__showToolbar = toolbarVisible  # 是否显示工具栏
        self.naviBar.setVisible(self.__showToolbar)
        layout = QtWidgets.QVBoxLayout(self.groupBox)
        layout.addWidget(self.naviBar)
        layout.addWidget(self.static_canvas)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        self._static_ax = self.static_canvas.figure.subplots()
        self.static_canvas.figure.subplots_adjust(bottom=0.15)

        self.__cid1 = self.static_canvas.mpl_connect("button_press_event", self.do_pressMouse)  # 支持曲线抓取
        self.__cid3 = self.static_canvas.mpl_connect("button_release_event", self.do_releaseMouse)  # 支持鼠标释放
        self.__cid4 = self.static_canvas.mpl_connect("motion_notify_event", self.do_moveMouse)  # 支持鼠标移动

        self.mouseIsPress = False
        self.pickStatus = False

        # self.get_gy_10_36()
        self.start_predict()

    def get_predict(self):
        self.counter+=1
        # a是列头字符串，b是列数据列表
        a = f"gy={self.structure_predict[3]:.1f},f={self.f:.4f}"
        b = self.spectrum_predict # 列表b

        # 选择活动的工作表
        ws = self.wb.active

        # 在第一列第一格写入字符串a
        ws.cell(row=1, column=self.counter, value=a)

        # 从第二行开始写入列表b中的数据
        for idx, value in enumerate(b, start=2):  # 从第二行开始
            ws.cell(row=idx, column=self.counter, value=value)

        # 指定保存的路径和文件名
        file_path = "./excels/finger_gy_10_36_v20144.xlsx"
        self.wb.save(file_path)  # 保存Excel文件

        # 关闭工作簿
        self.wb.close()

    def start_predict(self):
        """
        根据控制点进行插值  "nearest","zero","slinear","quadratic","cubic"
        :return:
        """
        global point
        self.start_predict_structure(point[0], point[1])

    def start_predict_structure(self, x, y):
        with torch.no_grad():
            d1 = self.lineEdit_d1.text()
            d2 = self.lineEdit_d2.text()
            gx = self.lineEdit_gx.text()
            gy = self.lineEdit_gy.text()
            s = [d2, d1, gx, gy]
            mask = [0, 0, 0, 0]
            for i, n in enumerate(s):
                try:
                    s[i] = float(n)
                    mask[i] = 1
                except ValueError:
                    s[i] = 0
                    mask[i] = 0
            v = [(x - 0.8) / 0.5]
            s = data_trans(s)
            reverse_mask = [1 if bit == 0 else 0 for bit in mask]
            input_s = [x * y for x, y in zip(s, mask)]
            input_s = [x if x != 0 else -1 for x in input_s]
            input_vs = v + input_s

            s = torch.tensor(s)
            input_vs = torch.tensor(input_vs).float()
            mask = torch.tensor(mask)
            reverse_mask = torch.tensor(reverse_mask)

            structure_out = i_model(input_vs)
            structure_out = structure_out.cpu()

            # 中途处理
            structure_2 = s.masked_fill(mask == 0, 0)
            structure_pre = structure_out.masked_fill(reverse_mask == 0, 0)
            structure_in = structure_2 + structure_pre

            valley_predict = f_model(structure_in)
            spectrum_predict = f2_model(structure_in)
            predict_structure = data_trans_inverse(structure_in).squeeze().tolist()
            predict_structure = [round(i, 1) for i in predict_structure]
            self._static_ax.cla()
            self._static_ax.set_xlim(0.8, 1.3)  # x 轴范围从 0 到 10
            self._static_ax.set_ylim(0, 1)
            # self._static_ax.set_xlabel('Frequency (THz)',
            #                            {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
            # self._static_ax.set_ylabel('Transmittance',
            #                            {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
            self._static_ax.set_xlabel(r'$f\,(\mathrm{THz})$',
                                       {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
            self._static_ax.set_ylabel('T',
                                       {'family': 'Arial', 'weight': 'normal', 'size': 16, 'style': 'italic'})  # 横坐标标签
            # self._static_ax.plot([i / 1000 * 0.5 + 0.8 for i in range(1001)], spectrum_predict.squeeze(),
            #                      label="Predicted Spectrum", color=(0x0A/255,0x6F/255,0xB4/255),linewidth=3, linestyle='--')
            # Set the font properties
            font_properties = font_manager.FontProperties(family='Arial', weight='normal', size=12)
            # Use the font properties in the legend
            self._static_ax.legend(fontsize=12, prop=font_properties)
            # plt.scatter(x=valley_predict * 0.5 + 0.8, y=0.4, s=400, c="blue")
            # plt.axvline(x=valley_predict * 0.5 + 0.8, color='r', linestyle='-', label="Customized pos", linewidth=2, )
            # min_value = min(spectrum_predict)
            # min_index = spectrum_predict.index(min_value)

            spectrum_predict = spectrum_predict.squeeze().tolist()
            self.spectrum_predict = spectrum_predict
            sigma = 2  # 高斯滤波器的标准差
            spectrum_predict_2 = gaussian_filter1d(spectrum_predict, sigma)
            min_index, min_value = self.get_valley(spectrum_predict_2)
            self._static_ax.scatter(x=(min_index / 1000) * 0.5 + 0.8, y=min_value, s=400, color=(0xEB/255,0x30/255,0x32/255))
            self._static_ax.axvline(x=(min_index / 1000) * 0.5 + 0.8, color=(0xEB/255,0x30/255,0x32/255), linestyle='-',
                                    label="Predicted Resonance",
                                    linewidth=2)
            self._static_ax.plot([i / 1000 * 0.5 + 0.8 for i in range(1001)], spectrum_predict_2,
                                 label="Predicted Spectrum", color=(0x0A / 255, 0x6F / 255, 0xB4 / 255), linewidth=3,
                                 linestyle='--')
            show_test = f"定制波谷：{(min_index / 1000) * 0.5 + 0.8:.4f} 预测结构:{predict_structure}"
            self.structure_predict = predict_structure
            print(show_test)

            self.f = (min_index / 1000) * 0.5 + 0.8;
            self.label_cr.setText(f"{(min_index / 1000) * 0.5 + 0.8:.4f}")
            self.label_d2.setText(f"{predict_structure[0]:.1f}")
            self.label_d1.setText(f"{predict_structure[1]:.1f}")
            self.label_gx.setText(f"{predict_structure[2]:.1f}")
            self.label_gy.setText(f"{predict_structure[3]:.1f}")

            self._static_ax.legend()
            self.static_canvas.draw()

    def clear_predict(self):
        self._static_ax.cla()
        self.static_canvas.draw()
        self.label_cr.setText("")
        self.label_d1.setText("")
        self.label_d2.setText("")
        self.label_gx.setText("")
        self.label_gy.setText("")
        self.lineEdit_d1.setText("")
        self.lineEdit_d2.setText("")
        self.lineEdit_gx.setText("")
        self.lineEdit_gy.setText("")

        self._static_ax.cla()
        self._static_ax.set_xlim(0.8, 1.3)  # x 轴范围从 0 到 10
        self._static_ax.set_ylim(0, 1)
        # self._static_ax.set_xlabel('Frequency (THz)',
        #                            {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        # self._static_ax.set_ylabel('Transmittance',
        #                            {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        self._static_ax.set_xlabel('Frequency (THz)',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        self._static_ax.set_ylabel('Transmittance',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        self.static_canvas.draw()

    def setToolbarVisible(self, isVisible=True):  # 是否显示工具栏
        self.__showToolbar = isVisible
        self.naviBar.setVisible(isVisible)

    def setDataHintVisible(self, isVisible=True):  # 是否显示坐标提示
        self.__showHint = isVisible
        self.__lastActtionHint.setVisible(isVisible)

    def do_pressMouse(self, event):
        if event.button == 1:
            global button_press_flag, update_nearlyPoint_flag
            button_press_flag = 1
            update_nearlyPoint_flag = 1
            self.point_move_show(event=event)

    def do_releaseMouse(self, event):  # 鼠标释放，释放抓取曲线
        global button_press_flag
        button_press_flag = 0

    def do_moveMouse(self, event):  # 鼠标移动，重绘抓取曲线
        if button_press_flag == 1:
            self.point_move_show(event=event)

    def point_move_show(self, **kw):
        """
        刷新图表上的点和曲线
        :param kw:
        :return:
        """
        global point

        if "event" in kw:
            self.update_point(kw["event"])
            # plt.scatter(point[0], 0.4, s=400, c="blue")
            # plt.draw()
        else:
            self._static_ax.scatter(point[0], 0.4, s=400, color="blue")
        self.start_predict()

    def update_point(self, event):
        """
        更新控制点
        :param event:
        :return:
        """
        global point, update_nearlyPoint_flag
        mouse_x = (event.xdata - 0.8) / 0.5
        mouse_y = event.ydata
        point_x = (point[0] - 0.8) / 0.5
        point_y = point[1]
        if event.xdata and event.ydata and 0 <= mouse_x <= 1 and 0 <= mouse_y <= 1:
            distance = math.sqrt(math.pow(mouse_x - point_x, 2) + math.pow(mouse_y - point_y, 2))
            if update_nearlyPoint_flag == 1:
                update_nearlyPoint_flag = 0
            if distance < 1:
                point = [event.xdata, event.ydata]

    def find_local_maximum(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_maximum = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] >= 0 and first_derivatives[i] <= 0:
                local_maximum.append(i)

        return local_maximum

    def find_local_minima(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_minima = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] <= 0 and first_derivatives[i] >= 0:
                local_minima.append(i)

        return local_minima

    def find_second_derivative_max_at_minima(self, data, minima_indices, h=0.0005):
        if not minima_indices:
            return None

        max_second_derivative_value = float('-inf')
        max_second_derivative_index = None

        # 计算每个局部极小值点的二阶导数
        for idx in minima_indices:
            if idx > 0 and idx < len(data) - 1:
                second_derivative = (data[idx + 1] - 2 * data[idx] + data[idx - 1]) / (h ** 2)
                if second_derivative > max_second_derivative_value:
                    max_second_derivative_value = second_derivative
                    max_second_derivative_index = idx

        return (max_second_derivative_index, data[max_second_derivative_index], max_second_derivative_value)

    def get_valley(self, spectrum_total):
        # self.valley_index = np.argmin(self.spectrum_total)
        # 找出所有局部极小值点
        minima_indices = self.find_local_minima(spectrum_total)
        # 在局部极小值点中找到二阶导数最大的点
        (valley_index, valley_value, max_second_derivative_value) = self.find_second_derivative_max_at_minima(
            spectrum_total, minima_indices)
        return valley_index,valley_value

    def get_gy_10_36(self):
        for gy in range(10,37):
            s = [6, 16, 6, gy]
            self.structure_predict = s.copy()
            s = data_trans(s)
            s = torch.tensor(s)
            spectrum_predict = f2_model(s)

            sigma = 2  # 高斯滤波器的标准差
            spectrum_predict = gaussian_filter1d(spectrum_predict.squeeze().tolist(), sigma)
            min_index, min_value = self.get_valley(spectrum_predict)
            self.f = (min_index / 1000) * 0.5 + 0.8;
            print(f"gy={gy}  f={self.f:.4f}\n")

            self.spectrum_predict = spectrum_predict
            self.get_predict()



if __name__ == '__main__':
    from PyQt5 import QtCore

    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    app = QtWidgets.QApplication(sys.argv)
    myshow = Mytest()
    myshow.show()
    sys.exit(app.exec_())
