import numpy as np
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QMessageBox

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, \
    NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

import sys

from openpyxl import load_workbook, Workbook
from scipy import interpolate, optimize

from structure_predict_form import Ui_Form

import os
import math
import torch
import numpy as np
from scipy.ndimage import gaussian_filter1d

import auxiliary_tools.parseUnit as parseUtils
from spectrum_unloaded_prediction_network.spectrum_prediction_network import ForwardNetMlp as forward2_design_net

args = parseUtils.MyParse(debug=False).args
args.device = "cpu"
f2_model = forward2_design_net().to(args.device)
forward2_tar = r"../data_space/checkpoints/spectrum-unloaded-prediction-network-checkpoint.pth.tar"
if os.path.exists(forward2_tar):
    print("load forward checkpoint...")
    f2_checkpoint = torch.load(forward2_tar)
    f2_model.load_state_dict(f2_checkpoint['state_dict'])
    print(f2_checkpoint['epoch'])
else:
    print("no found")


def data_trans_inverse(data):
    data = data.view(1, -1)
    scale = args.structure_range
    for i in range(0, data.shape[1]):
        data[0][i] = data[0][i] * (scale[i][1] - scale[i][0]) + scale[i][0]
    return data


def data_trans(data):
    scale = args.structure_range
    for i in range(0, data.__len__()):
        data[i] = (data[i] - scale[i][0]) / (scale[i][1] - scale[i][0])
    return data


global point, ims
global button_press_flag, nearly_point_index, update_nearlyPoint_flag
global true_spectrum, predict_spectrum, predict_structure
button_press_flag = 0
nearly_point_index = 0
update_nearlyPoint_flag = 0
global point  # 可移动的点，根据这些点的位置插值
point = [1.3, 0]


class Mytest(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(Mytest, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Structure_design")
        self.PredictButton.clicked.connect(self.start_predict)
        self.ClearButton.clicked.connect(self.clear_predict)
        self.GetButton.clicked.connect(self.get_predict)
        self.init()

    def init(self, toolbarVisible=True, showHint=False):
        self.counter = 1
        # 创建一个新的工作簿
        self.wb = openpyxl.Workbook()

        self.lineEdit_d1.setText('16')
        self.lineEdit_d2.setText('6')
        self.lineEdit_gx.setText('6')
        self.lineEdit_gy.setText('24')

        self.static_canvas = FigureCanvas(Figure())
        self.naviBar = NavigationToolbar(self.static_canvas, self)
        actList = self.naviBar.actions()
        count = len(actList)
        self.__lastActtionHint = actList[count - 1]
        self.__showHint = showHint  # 是否显示坐标提示
        self.__lastActtionHint.setVisible(self.__showHint)
        self.__showToolbar = toolbarVisible  # 是否显示工具栏
        self.naviBar.setVisible(self.__showToolbar)
        layout = QtWidgets.QVBoxLayout(self.groupBox)
        layout.addWidget(self.naviBar)
        layout.addWidget(self.static_canvas)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        self._static_ax = self.static_canvas.figure.subplots()

        self.__cid1 = self.static_canvas.mpl_connect("button_press_event", self.do_pressMouse)  # 支持曲线抓取
        self.__cid3 = self.static_canvas.mpl_connect("button_release_event", self.do_releaseMouse)  # 支持鼠标释放
        self.__cid4 = self.static_canvas.mpl_connect("motion_notify_event", self.do_moveMouse)  # 支持鼠标移动

        self.mouseIsPress = False
        self.pickStatus = False

        # self.get_10_36()
        self.start_predict()

    def get_predict_2(self):
        # 创建一个新的工作簿
        self.wb = openpyxl.Workbook()
        # 选择活动的工作表
        ws = self.wb.active

        # 在第一列第一格写入字符串a
        ws.cell(row=1, column=1, value="wave")
        ws.cell(row=1, column=2, value="spectrum")

        wavelength = np.arange(0.8, 1.3+0.0005, 0.0005).tolist()

        # 从第二行开始写入列表b中的数据
        for idx, value in enumerate(wavelength, start=2):  # 从第二行开始
            ws.cell(row=idx, column=1, value=value)
        for idx, value in enumerate(self.spectrum_predict, start=2):  # 从第二行开始
            ws.cell(row=idx, column=2, value=value)

        # 指定保存的路径和文件名
        filename = f"AI;d1={self.s[0]:.1f};d2={self.s[1]:.1f};gx={self.s[2]:.1f};gy={self.s[3]:.1f};f={self.f:.4f};i={self.i:.4f};q={self.q:.4f};w={self.w:.4f};"
        file_path = f"E:\\桌面\\703新增数据\\{filename}.xlsx"
        self.wb.save(file_path)  # 保存Excel文件

        # 关闭工作簿
        self.wb.close()

    def get_predict(self):
        filename = f"E:\\桌面\spnn_gai_d1.xlsx"

        # 如果文件存在，加载；否则新建
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # 指定保存的路径和文件名
        structure = f"d1={self.s[0]:.1f};d2={self.s[1]:.1f};gx={self.s[2]:.1f};gy={self.s[3]:.1f};"
        # 在第一列第一格写入字符串a
        ws.cell(row=self.counter, column=1, value=structure)
        ws.cell(row=self.counter, column=2, value=self.f)
        ws.cell(row=self.counter, column=3, value=self.i)
        ws.cell(row=self.counter, column=4, value=self.w)
        ws.cell(row=self.counter, column=5, value=self.q)
        self.counter += 1
        # 保存
        wb.save(filename)




    def get_10_36(self):
        for s in np.arange(6, 26.5, 0.5):
            s = [16, 6, 16, s]
            self.structure_predict = s.copy()
            s = [s[1], s[0], s[2], s[3]]
            s_nor = data_trans(s)
            s_nor = torch.tensor(s_nor)
            spectrum_out = f2_model(s_nor.float())
            spectrum_out = spectrum_out.cpu()

            spectrum_predict = spectrum_out.squeeze().tolist()
            sigma = 2  # 高斯滤波器的标准差
            spectrum_predict_smooth = gaussian_filter1d(spectrum_predict, sigma)
            min_index, min_value = self.get_valley(spectrum_predict_smooth)
            f0 = (min_index / 1000) * 0.5 + 0.8
            print(f"s={s}  f={f0:.4f}\n")

            rh = -1
            if rh < 0:
                [rh_pos, rh] = self.get_rh(spectrum_predict_smooth, f0)

            if rh > 0:
                hh = (rh + min_value) / 2

                x = [(i / 1000) * 0.5 + 0.8 for i in range(1001)]
                y = spectrum_predict_smooth
                # 创建三次样条插值函数
                f = interpolate.interp1d(x, y, kind="cubic")
                # 特定的y值
                y_target = hh
                # 定义新的函数 g(x) = f(x) - y_target
                def g(x):
                    return f(x) - y_target
                # 找到 g(x) = 0 的根
                x_found = []
                # 检查插值范围内的每个区间
                for i in range(len(x) - 1):
                    x0, x1 = x[i], x[i + 1]
                    if g(x0) * g(x1) < 0:  # 检查是否跨越y_target
                        root = optimize.root_scalar(g, bracket=[x0, x1]).root
                        x_found.append(root)
                print(f"y = {y_target} 时对应的 x 值有：{x_found}")

                if x_found.__len__() > 2:
                    def find_two_closest_values(lst, target):
                        # 计算每个元素与指定值的差值，并生成 (差值, 元素) 的元组列表
                        diffs = [(abs(x - target), x) for x in lst]
                        # 根据差值进行排序
                        diffs.sort()
                        # 选择差值最小的两个元素
                        closest_values = [diffs[0][1], diffs[1][1]]
                        return closest_values

                    # 找到与指定值最接近的两个值
                    x_found = find_two_closest_values(x_found, f0)

                if x_found.__len__() == 2:
                    f1 = min(x_found)
                    f2 = max(x_found)
                    fwhm = f2 - f1
                    q_factor = f0 / fwhm
                    self.excel_f = f0
                    self.excel_i = rh - min_value
                    self.excel_q = q_factor
                    self.excel_w = fwhm

            self.spectrum_predict = spectrum_predict
            # self.get_predict_10_36()

    def get_predict_10_36(self):
        self.counter+=1
        # 选择活动的工作表
        ws = self.wb.active
        # 在第一列第一格写入字符串a
        ws.cell(row=1, column=1, value="s")
        ws.cell(row=1, column=2, value="spnn-f0")
        ws.cell(row=1, column=3, value="spnn-h")
        ws.cell(row=1, column=4, value="spnn-w")
        ws.cell(row=1, column=5, value="spnn-q")

        # 在第一列第一格写入字符串a
        ws.cell(row=self.counter, column=1, value=f"d2={self.structure_predict[2]:.1f}")
        ws.cell(row=self.counter, column=2, value=f"{self.excel_f:.4f}")
        ws.cell(row=self.counter, column=3, value=f"{self.excel_i:.4f}")
        ws.cell(row=self.counter, column=4, value=f"{self.excel_w:.4f}")
        ws.cell(row=self.counter, column=5, value=f"{self.excel_q:.4f}")

        # 指定保存的路径和文件名
        file_path = r"E:\桌面\feature_gy_10_26.xlsx"
        self.wb.save(file_path)  # 保存Excel文件

        # 关闭工作簿
        self.wb.close()

    def start_predict(self):
        """
        根据控制点进行插值  "nearest","zero","slinear","quadratic","cubic"
        :return:
        """
        global point
        with torch.no_grad():
            d1 = self.lineEdit_d1.text()
            d2 = self.lineEdit_d2.text()
            gx = self.lineEdit_gx.text()
            gy = self.lineEdit_gy.text()
            rh = self.lineEdit_rh.text()
            s = [d2, d1, gx, gy]

            for i, n in enumerate(s):
                try:
                    s[i] = float(n)
                except ValueError:
                    QMessageBox.information(self, "信息提示框", "请输入正确的四个结构参数！")

            try:
                rh = float(rh)
            except ValueError:
                rh = -1
            self.s = s.copy()
            s_nor = data_trans(s)
            s_nor = torch.tensor(s_nor)
            spectrum_out = f2_model(s_nor)
            spectrum_out = spectrum_out.cpu()

            spectrum_predict = spectrum_out.squeeze().tolist()
            sigma = 2  # 高斯滤波器的标准差
            spectrum_predict_smooth = gaussian_filter1d(spectrum_predict, sigma)
            self.spectrum_predict = spectrum_predict_smooth
            min_index, min_value = self.get_valley(spectrum_predict_smooth)
            f0 = (min_index / 1000) * 0.5 + 0.8

            self._static_ax.cla()
            self._static_ax.set_xlim(0.8, 1.3)  # x 轴范围从 0 到 10
            self._static_ax.set_ylim(0, 1)
            self._static_ax.set_xlabel('Frequency (THz)',
                                       {'family': 'Arial', 'weight': 'normal', 'size': 20})  # 横坐标标签
            self._static_ax.set_ylabel('Transmittance',
                                       {'family': 'Arial', 'weight': 'normal', 'size': 20})  # 横坐标标签
            self._static_ax.plot([i / 1000 * 0.5 + 0.8 for i in range(1001)], spectrum_out.squeeze(),
                                 label="Predicted Spectrum",color="purple", linewidth=3, linestyle='-')
            self._static_ax.scatter(x=f0, y=min_value, s=300, c="blue")

            # self._static_ax.scatter(x=1.3, y=point[1], s=400, c="r")
            # self._static_ax.axhline(y=point[1], color='r', linestyle='--')
            self._static_ax.scatter(x=point[0], y=point[1], s=100, c="r")
            self._static_ax.plot([point[0] - 0.03, point[0] + 0.03], [point[1], point[1]], 'b--')  # 横线
            self._static_ax.plot([point[0], point[0]], [point[1] - 0.08, point[1] + 0.08], 'b--')  # 竖线
            self.label_x.setText(f"{point[0]:.4f}")
            self.label_y.setText(f"{point[1]:.4f}")
            if point[1]>0.3:
                rh = point[1]
            else:
                rh = -1

            if rh < 0:
                [rh_pos, rh] = self.get_rh(spectrum_predict_smooth, f0)

            if rh > 0:
                hh = (rh + min_value) / 2
                self._static_ax.axhline(y=rh, color='b', linestyle='--')

                x = [(i / 1000) * 0.5 + 0.8 for i in range(1001)]
                y = spectrum_predict_smooth
                # 创建三次样条插值函数
                f = interpolate.interp1d(x, y, kind="cubic")
                # 特定的y值
                y_target = hh
                # 定义新的函数 g(x) = f(x) - y_target
                def g(x):
                    return f(x) - y_target
                # 找到 g(x) = 0 的根
                x_found = []
                # 检查插值范围内的每个区间
                for i in range(len(x) - 1):
                    x0, x1 = x[i], x[i + 1]
                    if g(x0) * g(x1) < 0:  # 检查是否跨越y_target
                        root = optimize.root_scalar(g, bracket=[x0, x1]).root
                        x_found.append(root)
                print(f"y = {y_target} 时对应的 x 值有：{x_found}")

                if x_found.__len__() > 2:
                    def find_two_closest_values(lst, target):
                        # 计算每个元素与指定值的差值，并生成 (差值, 元素) 的元组列表
                        diffs = [(abs(x - target), x) for x in lst]
                        # 根据差值进行排序
                        diffs.sort()
                        # 选择差值最小的两个元素
                        closest_values = [diffs[0][1], diffs[1][1]]
                        return closest_values

                    # 找到与指定值最接近的两个值
                    x_found = find_two_closest_values(x_found, f0)

                if x_found.__len__() == 2:
                    f1 = min(x_found)
                    f2 = max(x_found)
                    fwhm = f2 - f1
                    q_factor = f0 / fwhm
                    self.label_f0.setText(f"{f0:.4f}")
                    self.label_i.setText(f"{rh - min_value:.4f}")
                    self.label_w.setText(f"{fwhm:.4f}")
                    self.label_q.setText(f"{q_factor:.4f}")
                    self._static_ax.plot([f1, f2], [hh, hh], 'p--')  # 横线

                if x_found.__len__() < 2:
                    QMessageBox.information(self, "信息提示框", "无法求解得到对应的f1和f2")
                self.f = f0
                self.i = rh - min_value
                self.q = q_factor
                self.w = fwhm

            self._static_ax.legend()
            self.static_canvas.draw()

    def clear_predict(self):
        # self._static_ax.cla()
        # self.static_canvas.draw()
        # self.label_cr.setText("")
        # self.label_d1.setText("")
        # self.label_d2.setText("")
        # self.label_gx.setText("")
        # self.label_gy.setText("")
        # self.lineEdit_d1.setText("")
        # self.lineEdit_d2.setText("")
        # self.lineEdit_gx.setText("")
        # self.lineEdit_gy.setText("")
        #
        # self._static_ax.cla()
        # self._static_ax.set_xlim(0.8, 1.3)  # x 轴范围从 0 到 10
        # self._static_ax.set_ylim(0, 1)
        # self._static_ax.set_xlabel('Frequency (THz)',
        #                            {'family': 'Arial', 'weight': 'normal', 'size': 20})  # 横坐标标签
        # self._static_ax.set_ylabel('Transmittance',
        #                            {'family': 'Arial', 'weight': 'normal', 'size': 20})  # 横坐标标签
        self.static_canvas.draw()

    def setToolbarVisible(self, isVisible=True):  # 是否显示工具栏
        self.__showToolbar = isVisible
        self.naviBar.setVisible(isVisible)

    def setDataHintVisible(self, isVisible=True):  # 是否显示坐标提示
        self.__showHint = isVisible
        self.__lastActtionHint.setVisible(isVisible)

    def do_pressMouse(self, event):
        if event.button == 1:
            global button_press_flag, update_nearlyPoint_flag
            button_press_flag = 1
            update_nearlyPoint_flag = 1
            self.point_move_show(event=event)

    def do_releaseMouse(self, event):  # 鼠标释放，释放抓取曲线
        global button_press_flag
        button_press_flag = 0

    def do_moveMouse(self, event):  # 鼠标移动，重绘抓取曲线
        if button_press_flag == 1:
            self.point_move_show(event=event)

    def point_move_show(self, **kw):
        """
        刷新图表上的点和曲线
        :param kw:
        :return:
        """
        global point

        if "event" in kw:
            self.update_point(kw["event"])
        self.start_predict()

    def update_point(self, event):
        """
        更新控制点
        :param event:
        :return:
        """
        global point, update_nearlyPoint_flag
        mouse_x = (event.xdata - 0.8) / 0.5
        mouse_y = event.ydata
        point_x = (point[0] - 0.8) / 0.5
        point_y = point[1]
        if event.xdata and event.ydata and 0 <= mouse_x <= 1 and 0 <= mouse_y <= 1:
            distance = math.sqrt(math.pow(mouse_x - point_x, 2) + math.pow(mouse_y - point_y, 2))
            if update_nearlyPoint_flag == 1:
                update_nearlyPoint_flag = 0
            if distance < 1:
                point = [event.xdata, event.ydata]

    def find_local_maximum(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_maximum = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] >= 0 and first_derivatives[i] <= 0:
                local_maximum.append(i)

        return local_maximum

    def find_local_minima(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_minima = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] <= 0 and first_derivatives[i] >= 0:
                local_minima.append(i)

        return local_minima

    def find_second_derivative_max_at_minima(self, data, minima_indices, h=0.0005):
        if not minima_indices:
            return None

        max_second_derivative_value = float('-inf')
        max_second_derivative_index = None

        # 计算每个局部极小值点的二阶导数
        for idx in minima_indices:
            if idx > 0 and idx < len(data) - 1:
                second_derivative = (data[idx + 1] - 2 * data[idx] + data[idx - 1]) / (h ** 2)
                if second_derivative > max_second_derivative_value:
                    max_second_derivative_value = second_derivative
                    max_second_derivative_index = idx

        return (max_second_derivative_index, data[max_second_derivative_index], max_second_derivative_value)

    def get_valley(self, spectrum_total):
        # self.valley_index = np.argmin(self.spectrum_total)
        # 找出所有局部极小值点
        minima_indices = self.find_local_minima(spectrum_total)
        # 在局部极小值点中找到二阶导数最大的点
        (valley_index, valley_value, max_second_derivative_value) = self.find_second_derivative_max_at_minima(
            spectrum_total, minima_indices)
        return valley_index, valley_value

    def get_rh(self, spectrum_total, f0):
        # self.valley_index = np.argmin(self.spectrum_total)
        # 找出所有局部极大值点和边值
        maximum_indices = self.find_local_maximum(spectrum_total)
        maximum_pos = [i / 1000 * 0.5 + 0.8 for i in maximum_indices]
        maximum_pos.append(0.8)
        maximum_pos.append(1.3)

        # def find_two_closest_values(lst, target):
        #     # 计算每个元素与指定值的差值，并生成 (差值, 元素) 的元组列表
        #     diffs = [(abs(x - target), x) for x in lst]
        #     # 根据差值进行排序
        #     diffs.sort()
        #     # 选择差值最小的两个元素
        #     closest_values = [diffs[0][1], diffs[1][1]]
        #     return closest_values
        #
        # # 找到与指定值最接近的两个值
        # maximum_pos = find_two_closest_values(maximum_pos, f0)
        # maximum_indices = [spectrum_total[int((maximum_pos[0] - 0.8) * 2000)],spectrum_total[int((maximum_pos[1] - 0.8) * 2000)]]
        # rh = min(maximum_indices)
        # rh_index = spectrum_total.tolist().index(rh)

        def find_one_closest_values(lst, target):
            # 计算每个元素与指定值的差值，并生成 (差值, 元素) 的元组列表
            diffs = [(abs(x - target), x) for x in lst]
            # 根据差值进行排序
            diffs.sort()
            # 选择差值最小的两个元素
            closest_values = diffs[0][1]
            return closest_values

        # 找到与指定值最接近的两个值
        maximum_pos = find_one_closest_values(maximum_pos, f0)
        rh = spectrum_total[int((maximum_pos - 0.8) * 2000)]
        rh_index = spectrum_total.tolist().index(rh)
        return [rh_index / 1000 * 0.5 + 0.8 , rh]

if __name__ == '__main__':
    from PyQt5 import QtCore

    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    app = QtWidgets.QApplication(sys.argv)
    myshow = Mytest()
    myshow.show()
    sys.exit(app.exec_())
