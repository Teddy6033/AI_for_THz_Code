import numpy as np
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas, NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

import sys

from scipy.ndimage import gaussian_filter1d

from RI_S_Predict_form import Ui_Form

import os
import math
import torch
import numpy as np

import auxiliary_tools.parseUnit as parseUtils
from spectrum_unloaded_prediction_network.spectrum_prediction_network import ForwardNetMlp as forward2_design_net

args = parseUtils.MyParse(debug=False).args
args.device = "cpu"
f2_model = forward2_design_net().to(args.device)
loaded_f2_model = forward2_design_net().to(args.device)

forward2_tar = r"../data_space/checkpoints/spectrum-unloaded-prediction-network-checkpoint.pth.tar"
if os.path.exists(forward2_tar):
    print("load forward checkpoint...")
    f2_checkpoint = torch.load(forward2_tar)
    f2_model.load_state_dict(f2_checkpoint['state_dict'])
    print(f2_checkpoint['epoch'])
else:
    print("no found")

loaded_forward2_tar = r"../data_space/checkpoints/spectrum-loaded-prediction-network-checkpoint.pth.tar"
if os.path.exists(loaded_forward2_tar):
    print("load forward checkpoint...")
    loaded_f2_checkpoint = torch.load(loaded_forward2_tar)
    loaded_f2_model.load_state_dict(loaded_f2_checkpoint['state_dict'])
    print(loaded_f2_checkpoint['epoch'])
else:
    print("no found")

def data_trans_inverse(data):
    data = data.view(1, -1)
    scale = args.structure_range
    for i in range(0, data.shape[1]):
        data[0][i] = data[0][i] * (scale[i][1] - scale[i][0]) + scale[i][0]
    return data

def data_trans(data):
    scale = args.structure_range
    for i in range(0, data.__len__()):
        data[i] = (data[i] - scale[i][0]) / (scale[i][1] - scale[i][0])
    return data

class Mytest(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(Mytest, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("RI_S_predict")
        self.StartButton.clicked.connect(self.start_predict)
        self.ClearButton.clicked.connect(self.clear_predict)
        self.GetButton.clicked.connect(self.get_predict)
        self.init()

    def init(self):
        self.lineEdit_d1.setText('16')
        self.lineEdit_d2.setText('6')
        self.lineEdit_gx.setText('6')
        self.lineEdit_gy.setText('24')

        self.static_canvas = FigureCanvas(Figure())
        layout = QtWidgets.QVBoxLayout(self.groupBox)
        layout.addWidget(self.static_canvas)
        self._static_ax = self.static_canvas.figure.subplots()
        self.static_canvas.figure.subplots_adjust(bottom=0.15)

    def start_predict(self):
        d1 = self.lineEdit_d1.text()
        d2 = self.lineEdit_d2.text()
        gx = self.lineEdit_gx.text()
        gy = self.lineEdit_gy.text()
        s = [d2,d1,gx,gy]
        s = [float(i) for i in s]
        self.d2 = s[0]
        self.d1 = s[1]
        self.gx = s[2]
        self.gy = s[3]
        s = data_trans(s)
        structure_in = torch.tensor(s)
        spectrum_predict = f2_model(structure_in)
        loaded_spectrum_predict = loaded_f2_model(structure_in)
        predict_structure = data_trans_inverse(structure_in).squeeze().tolist()
        predict_structure = [round(i, 2) for i in predict_structure]

        # self._static_ax.xlabel('Frequency(THz)', {'family': 'Times New Roman', 'weight': 'normal', 'size': 20})
        # self._static_ax.ylabel('Transmissivity', {'family': 'Times New Roman', 'weight': 'normal', 'size': 20})
        self._static_ax.cla()
        self._static_ax.plot([i / 1000 * 0.5 + 0.8 for i in range(1001)], spectrum_predict.squeeze().tolist(),
                 label="Bare spectrum", color=(0x0A/255,0x6F/255,0xB4/255), linewidth=3, linestyle='--')
        self._static_ax.plot([i / 1000 * 0.5 + 0.8 for i in range(1001)], loaded_spectrum_predict.squeeze().tolist(),
                 label="Loaded spectrum", color=(0xEB/255,0x30/255,0x32/255), linewidth=3, linestyle='--')
        # self._static_ax.set_xlabel('Frequency (THz)', {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        # self._static_ax.set_ylabel('Transmittance', {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签

        self._static_ax.set_xlabel(r'$f\,(\mathrm{THz})$',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16})  # 横坐标标签
        self._static_ax.set_ylabel('T',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16, 'style': 'italic'})  # 横坐标标签

        sigma = 2  # 高斯滤波器的标准差

        spectrum_predict = spectrum_predict.squeeze().tolist()
        spectrum_predict_2 = gaussian_filter1d(spectrum_predict, sigma)
        self.unloed_spectrum = spectrum_predict_2.tolist()
        unloaded_min_index, unloaded_min_value = self.get_valley(spectrum_predict_2)
        self._static_ax.axvline(x=(unloaded_min_index / 1000) * 0.5 + 0.8, color=(0x0A/255,0x6F/255,0xB4/255), linestyle='-', label="Bare resonance",
                    linewidth=3, )

        loaded_spectrum_predict = loaded_spectrum_predict.squeeze().tolist()
        loaded_spectrum_predict_2 = gaussian_filter1d(loaded_spectrum_predict, sigma)
        self.loed_spectrum = loaded_spectrum_predict_2.tolist()
        loaded_min_index, loaded_min_value = self.get_valley(loaded_spectrum_predict_2)
        self._static_ax.axvline(x=(loaded_min_index / 1000) * 0.5 + 0.8, color=(0xEB/255,0x30/255,0x32/255), linestyle='-', label="Loaded resonance",
                    linewidth=3, )

        unloaded_resonance = unloaded_min_index / 1000 * 0.5 + 0.8
        loaded_resonance = loaded_min_index / 1000 * 0.5 + 0.8
        unloaded_resonance = round(unloaded_resonance, 4)
        loaded_resonance = round(loaded_resonance, 4)
        S = (loaded_resonance - unloaded_resonance) / (-0.7) * 1000 / 1.5
        Sn = S / ((loaded_resonance + unloaded_resonance) / 2) / 1.5
        show_test = f"预测结构:{predict_structure} 有载波谷：{loaded_resonance:.4f} 无载波谷：{unloaded_resonance:.4f}"
        print(show_test)
        self.label_ur.setText(f" {unloaded_resonance:.4f}")
        self.label_lr.setText(f" {loaded_resonance:.4f}")
        self.label_S.setText(f"  {S:.1f}")
        # self.label_Sn.setText(f"  {Sn:.1f}")
        self.ur = unloaded_resonance
        self.lr = loaded_resonance
        self.S = S
        self.Sn = Sn
        # show_test = (
        #     f"d1:{predict_structure[0]}um  d2:{predict_structure[1]}um  gx:{predict_structure[2]}um  gy:{predict_structure[3]}um"
        #     f"\nur:{unloaded_resonance:.4f}THz lr:{loaded_resonance:.4f}THz S:{S:.1f}GHz/RIU Sn:{Sn:.1f}/RIU/um")
        # self._static_ax.title(show_test, fontsize=14, fontstyle='italic')
        self._static_ax.legend()
        self.static_canvas.draw()

    def get_predict(self):
        # 创建一个新的工作簿
        self.wb = openpyxl.Workbook()
        # 选择活动的工作表
        ws = self.wb.active

        # 在第一列第一格写入字符串a
        ws.cell(row=1, column=1, value="wave")
        ws.cell(row=1, column=2, value="unload")
        ws.cell(row=1, column=3, value="load")

        wavelength = np.arange(0.8, 1.3+0.0005, 0.0005).tolist()

        # 从第二行开始写入列表b中的数据
        for idx, value in enumerate(wavelength, start=2):  # 从第二行开始
            ws.cell(row=idx, column=1, value=value)
        for idx, value in enumerate(self.unloed_spectrum, start=2):  # 从第二行开始
            ws.cell(row=idx, column=2, value=value)
        for idx, value in enumerate(self.loed_spectrum, start=2):  # 从第二行开始
            ws.cell(row=idx, column=3, value=value)

        # 指定保存的路径和文件名
        filename = f"d1={self.d1:.1f},d2={self.d2:.1f},gx={self.gx:.1f},gy={self.gy:.1f},ur={self.ur:.4f},lr={self.lr:.4f},S={self.S:.1f},Sn={self.Sn:.1f}"
        file_path = f"./excels/{filename}.xlsx"
        self.wb.save(file_path)  # 保存Excel文件

        # 关闭工作簿
        self.wb.close()

    def clear_predict(self):
        self._static_ax.cla()
        self.static_canvas.draw()
        self.label_ur.setText("")
        self.label_lr.setText("")
        self.label_S.setText("")
        # self.label_Sn.setText("")
        self.lineEdit_d1.setText("")
        self.lineEdit_d2.setText("")
        self.lineEdit_gx.setText("")
        self.lineEdit_gy.setText("")

        self._static_ax.cla()
        self._static_ax.set_xlim(0.8, 1.3)  # x 轴范围从 0 到 10
        self._static_ax.set_ylim(0, 1)
        self._static_ax.set_xlabel('f(THz)',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16, 'style': 'italic'})  # 横坐标标签
        self._static_ax.set_ylabel('T',
                                   {'family': 'Arial', 'weight': 'normal', 'size': 16, 'style': 'italic'})  # 横坐标标签
        self.static_canvas.draw()

    def find_local_maximum(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_maximum = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] >= 0 and first_derivatives[i] <= 0:
                local_maximum.append(i)

        return local_maximum

    def find_local_minima(self, data, h=0.0005):
        n = len(data)
        if n < 3:
            return []  # 数据点不足，无法找到局部极小值点

        local_minima = []

        # 计算一阶导数
        first_derivatives = []
        for i in range(1, n - 1):
            first_derivative = (data[i + 1] - data[i - 1]) / (2 * h)
            first_derivatives.append(first_derivative)

        # 寻找局部极小值点
        for i in range(1, n - 2):
            if first_derivatives[i - 1] <= 0 and first_derivatives[i] >= 0:
                local_minima.append(i)

        return local_minima

    def find_second_derivative_max_at_minima(self, data, minima_indices, h=0.0005):
        if not minima_indices:
            return None

        max_second_derivative_value = float('-inf')
        max_second_derivative_index = None

        # 计算每个局部极小值点的二阶导数
        for idx in minima_indices:
            if idx > 0 and idx < len(data) - 1:
                second_derivative = (data[idx + 1] - 2 * data[idx] + data[idx - 1]) / (h ** 2)
                if second_derivative > max_second_derivative_value:
                    max_second_derivative_value = second_derivative
                    max_second_derivative_index = idx

        return (max_second_derivative_index, data[max_second_derivative_index], max_second_derivative_value)

    def get_valley(self, spectrum_total):
        # self.valley_index = np.argmin(self.spectrum_total)
        # 找出所有局部极小值点
        minima_indices = self.find_local_minima(spectrum_total)
        # 在局部极小值点中找到二阶导数最大的点
        (valley_index, valley_value, max_second_derivative_value) = self.find_second_derivative_max_at_minima(
            spectrum_total, minima_indices)
        return valley_index, valley_value



if __name__ == '__main__':
    from PyQt5 import QtCore

    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    app = QtWidgets.QApplication(sys.argv)
    myshow = Mytest()
    myshow.show()
    sys.exit(app.exec_())
