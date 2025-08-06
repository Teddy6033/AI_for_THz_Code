import os
import sys
import numpy as np
import openpyxl
import torch
import torch.nn as nn
from sklearn.metrics import r2_score, mean_absolute_error
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt

import auxiliary_tools.parseUnit as parseUtils
from bidirectional_forward_design_network.forward_design_network import ForwardNetMlp as F_Net
from auxiliary_tools.dataSet_inverse import DataSet
from inverse_design_network import InverseNet as I_Net

args = parseUtils.MyParse(debug=False).args
dataset = torch.load(r"..\data_space\datasets_tar\unloaded_dataset_inverse_for_technology.pth")
dataset_train = dataset['dataset_train']
dataset_val = dataset['dataset_val']

train_loader = DataLoader(dataset=dataset_train, batch_size=64, shuffle=True)
val_loader = DataLoader(dataset=dataset_val, batch_size=64, shuffle=True)

f_model = F_Net().to(args.device)
forward_tar = r"../data_space/checkpoints/bidirectional-forward-design-network-checkpoint.pth.tar"

if os.path.exists(forward_tar):
    print("load forward checkpoint...")
    checkpoint = torch.load(forward_tar)
    f_model.load_state_dict(checkpoint['state_dict'])
    print(checkpoint['epoch'])
else:
    print("no found")

# 加载逆向模型
i_model = I_Net().to(args.device)
inverse_tar = r"../data_space/checkpoints/bidirectional-inverse-design-network-checkpoint.pth.tar"
if os.path.exists(inverse_tar):
    print("load inverse checkpoint...")
    checkpoint = torch.load(inverse_tar)
    i_model.load_state_dict(checkpoint['state_dict'])
    print(checkpoint['epoch'])
else:
    print("no found")


def evaluate_model(loader, f_model, i_model, device):
    """计算模型在给定数据加载器上的MSE、R²、MAE和RMSE"""
    f_model.eval()
    i_model.eval()
    all_preds = []
    all_labels = []
    running_loss = 0.0  # 用于累加损失

    with torch.no_grad():
        for index, (input_vs, valley_pos, real_structure, mask,
                    reverse_mask, structure_raw, spectrum_total) in enumerate(loader):
            input_vs = input_vs.to(device)
            valley_pos = valley_pos.to(device)
            real_structure = real_structure.to(device)
            mask = mask.to(device)
            reverse_mask = reverse_mask.to(device)

            # 逆向传播
            structure_out = i_model(input_vs)
            # 中途处理
            structure_1 = structure_out.masked_fill(mask == 0, 0)
            structure_2 = real_structure.masked_fill(mask == 0, 0)
            structure_pre = structure_out.masked_fill(reverse_mask == 0, 0)
            structure_in = structure_2 + structure_pre

            # 正向传播
            valley_out = f_model(structure_in)
            loss_valley = nn.MSELoss()(valley_pos, valley_out)
            # 计算损失
            loss = loss_valley

            # 保存预测值和标签
            all_preds.append(valley_out.cpu().numpy())
            all_labels.append(valley_pos.cpu().numpy())

            running_loss += loss.item() * real_structure.size(0)  # 累加总损失

            # 更新进度
            progress = (index + 1) / len(loader) * 100
            sys.stdout.write(f'\rCurrent Progress: {progress:.2f}%')
            sys.stdout.flush()

    # 合并所有预测和标签
    preds = np.concatenate(all_preds, axis=0)
    labels = np.concatenate(all_labels, axis=0)

    # 计算平均MSE
    mse = running_loss / len(loader.dataset)  # 计算平均MSE
    # 计算RMSE
    rmse = np.sqrt(mse)  # 计算RMSE
    # 计算R²和MAE
    r2 = r2_score(labels, preds)
    mae = mean_absolute_error(labels, preds)

    return mse, rmse, r2, mae, preds, labels  # 返回预测值和标签


# 训练集评估
train_loss, train_rmse, train_r2, train_mae, train_preds, train_labels = evaluate_model(train_loader, f_model, i_model,
                                                                                        args.device)
print(f"\nTrain MSE: {train_loss:.7f}")  # MSE
print(f"Train RMSE: {train_rmse:.7f}")  # RMSE
print(f"Train R²: {train_r2:.7f}")  # R²
print(f"Train MAE: {train_mae:.7f}")  # MAE

# 验证集评估
val_loss, val_rmse, val_r2, val_mae, val_preds, val_labels = evaluate_model(val_loader, f_model, i_model, args.device)
print(f"\nValidation MSE: {val_loss:.7f}")  # MSE
print(f"Validation RMSE: {val_rmse:.7f}")  # RMSE
print(f"Validation R²: {val_r2:.7f}")  # R²
print(f"Validation MAE: {val_mae:.7f}")  # MAE

# 设置随机种子并抽取1000个样本
np.random.seed(1314)
train_indices = np.random.choice(len(train_preds), 1000, replace=False)
val_indices = np.random.choice(len(val_preds), 1000, replace=False)

# 获取抽样后的数据
train_preds_sample = train_preds[train_indices]
train_labels_sample = train_labels[train_indices]
val_preds_sample = val_preds[val_indices]
val_labels_sample = val_labels[val_indices]

# 绘制散点图
plt.figure(figsize=(10, 6))

# # 创建一个新的工作簿
# wb = openpyxl.Workbook()
# # 选择活动的工作表
# ws = wb.active
# # 在第一列第一格写入字符串a
# ws.cell(row=1, column=1, value="pred")
# ws.cell(row=1, column=2, value="label")
# val_preds_sample_list = val_preds_sample.tolist()
# val_labels_sample_list = val_labels_sample.tolist()
# # 从第二行开始写入列表b中的数据
# for idx, value in enumerate(val_preds_sample, start=2):  # 从第二行开始
#     ws.cell(row=idx, column=1, value=float(value))
# for idx, value in enumerate(val_labels_sample, start=2):  # 从第二行开始
#     ws.cell(row=idx, column=2, value=float(value))
# # 指定保存的路径和文件名
# filename = f"SINN_pred_vs_label"
# file_path = f"E:/桌面/{filename}.xlsx"
# wb.save(file_path)  # 保存Excel文件
# # 关闭工作簿
# wb.close()

# 验证集散点图 - 实心蓝色正方形
plt.xlim(0, 1)
plt.ylim(0, 1)
plt.scatter(val_preds_sample, val_labels_sample, color='blue', marker='s', label='Validation Set', alpha=0.4, s=50)
# 训练集散点图 - 红色圆点
# plt.scatter(train_preds_sample, train_labels_sample, color='red', marker='o', label='Training Set', alpha=0.4, s=50)

# 图例和标签
plt.xlabel('Predicted Value')
plt.ylabel('True Value')
plt.legend()
plt.title('Predicted vs True Values for Training and Validation Sets')
plt.show()
